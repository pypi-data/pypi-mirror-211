#################################
# class for handling excel
#################################

# these classes can be re-used to place data frames or images into excel sheets

import openpyxl
import re
import os

"""
# https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/
"""

# This class was written 90% by ChatGPT

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class ExcelStyles:
    def __init__(self, header_font='Calibri', header_size=12, header_bold=True,
                 index_font='Calibri', index_size=11, data_font='Calibri', data_size=11):
        self.header_font = Font(name=header_font, size=header_size, bold=header_bold)
        self.index_font = Font(name=index_font, size=index_size)
        self.data_font = Font(name=data_font, size=data_size)
        
        self.header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def style_sheet(self, sheet, df, start_row=1, start_col=1, add_rownames=True):
        num_rows, num_cols = df.shape
        
        # Set style for header cells
        for col in range(num_cols):
            cell = sheet.cell(row=start_row, column=start_col+col+1 if add_rownames else start_col+col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        # Set style for index cells
        if add_rownames:
            for row in range(num_rows):
                cell = sheet.cell(row=start_row+row+1, column=start_col)
                cell.font = self.index_font
                cell.border = self.thin_border
        
        # Set style for data cells
        for row in range(num_rows):
            for col in range(num_cols):
                cell = sheet.cell(row=start_row+row+1, column=start_col+col+1 if add_rownames else start_col + col)
                cell.font = self.data_font
                cell.border = self.thin_border

    def adjust_column_widths(self, ws, df, start_row, start_col, add_rownames=True):
        for i, column in enumerate(df.columns):
            col_letter = openpyxl.utils.get_column_letter(start_col + i + 1 if add_rownames else start_col + i)
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(column)
            )
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2
        
    def convert_to_scientific_notation(self, df, col_name):
        # create a copy of the dataframe to avoid modifying the original
        df_new = df.copy()

        # convert the column to scientific notation strings
        df_new[col_name] = df_new[col_name].apply(lambda x: '{:.2E}'.format(x))

        return df_new
    
    def style_column(self, file_path, sheet_name, df, start_row, start_col, col, bold=False, italic=False, font_color=None, fill_color=None, add_rownames=True):
        workbook = openpyxl.load_workbook(filename=file_path)
        sheet = workbook[sheet_name]

        if isinstance(col, int):  # Check if col is already a column index
            col_index = col
        elif isinstance(col, str):  # Check if col is a column name
            col_index = df.columns.tolist().index(col)
        else:
            raise ValueError("Invalid column identifier. Must be an integer index or a string name.")

        for row_num in range(start_row, start_row + len(df) + 1): # +1 assuming header is always attached
            cell = sheet.cell(row=row_num, column=start_col + col_index if add_rownames else start_col + col_index - 1)
            font = cell.font
            if bold:
                font = Font(name=font.name, size=font.size, bold=True)
            if italic:
                font = Font(name=font.name, size=font.size, italic=True)
            if font_color:
                font = Font(name=font.name, size=font.size, color=font_color)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.font = font

        workbook.save(filename=file_path)

# This class was written by ChatGPT in form of different classes
# I composed them to one class

class ExcelWriter:
    @staticmethod
    def write_dataframe(file_path, sheet_name, df, start_row, start_col, sheet_title=None, add_rownames=True):
        
        # Load the workbook
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(filename=file_path)
        else:
            workbook = openpyxl.Workbook()
        
        # Select the worksheet
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        ws = ExcelWriter.create_or_get_sheet(workbook, sheet_name)

        if sheet_title:
            ws.title = sheet_title

        # write the column headers to the worksheet
        for col, header in enumerate(df.columns):
            ws.cell(row=start_row, column=start_col+col+1 if add_rownames else start_col+col, value=header)

        # write the row indexes to the worksheet if required
        if add_rownames:
            for row, index in enumerate(df.index):
                ws.cell(row=start_row+row+1, column=start_col, value=index)

        # write the data from the df to the worksheet
        for row, series in enumerate(df.itertuples(index=False)):
            for col, value in enumerate(series):
                ws.cell(row=start_row+row+1, column=start_col + col + 1 if add_rownames else start_col+col, value=value)

        # Style the sheet
        styles = ExcelStyles()
        styles.style_sheet(ws, df, start_row=start_row, start_col=start_col, add_rownames=add_rownames)
        styles.adjust_column_widths(ws, df, start_row=start_row, start_col=start_col, add_rownames=add_rownames)

        if not os.path.exists(file_path):
            # delete empty first sheet
            workbook.remove_sheet(workbook._sheets[0])

        # Save the changes to the workbook
        (workbook if os.path.exists(file_path) else ExcelWriter.delete_first_empty_sheet(workbook)).save(filename=file_path)
    
    @staticmethod
    def write_image(file_path, sheet_name, image_path, start_row, start_col, sheet_title=None):
        # Load the workbook
        # Load the workbook
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(filename=file_path)
        else:
            workbook = openpyxl.Workbook()
        
        # Select the worksheet
        worksheet = ExcelWriter.create_or_get_sheet(workbook, sheet_name)

        if sheet_title:
            worksheet.title = sheet_title
        
        # Add the image to the worksheet
        img = openpyxl.drawing.image.Image(image_path)
        worksheet.add_image(img, f"{openpyxl.utils.get_column_letter(start_col)}{start_row}")

        # Save the changes to the workbook
        (workbook if os.path.exists(file_path) else ExcelWriter.delete_first_empty_sheet(workbook)).save(filename=file_path)

    @staticmethod
    def excel_to_rc(cell_ref):
        # Split the cell reference into its column and row parts
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if not match:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        
        col_ref, row_ref = match.groups()
        
        # Convert the column reference to a number
        col_num = 0
        for char in col_ref:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
        # Convert the row reference to a number
        row_num = int(row_ref)
    
        return row_num, col_num

    @staticmethod
    def rc_to_excel(row, col):
        # Convert the column number to a letter code
        col_ref = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            col_ref = chr(65 + remainder) + col_ref

        # combine the column and row references
        return col_ref + str(row)

    @staticmethod
    def create_or_get_sheet(wb, sheet_name):
        
        if sheet_name in wb.sheetnames:
            # sheet already exists, so return it
            return wb[sheet_name]
        else:
            # sheet doesn't exist, so create it
            return wb.create_sheet(sheet_name)

    @staticmethod
    def delete_first_empty_sheet( wb):
        first_sheet = wb._sheets[0]
        if 'Sheet' == first_sheet.title and [] == list(first_sheet.values):
            wb.remove_sheet(wb._sheets[0])
        return wb