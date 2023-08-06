# coding=utf-8
"""
(C) FHCS GmbH

Generic ViledaUtils class implementing static methods to use for MS Excel import / export and API communication
"""

import logging
from datetime import datetime, timedelta
from os import walk
from traceback import format_exc
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from .importsheetparser import ImportSheetParser
from .viledaexceptions import ImportCellParserException, ImportSheetParserException


class ViledaUtils:
    """
    Static class containig various methods to use for import / export, API communication, convertations
    """

    def __init__(self):
        pass

    @staticmethod
    def loadexcels(xlsdir: str,
                   tolog: logging.Logger,
                   pfile: Optional[str] = None,
                   loglist: Optional[List[Dict[str, Any]]] = None,
                   logdebug: bool = False) -> List[Dict[str, Any]]:
        """
        Loads all excels from given folder, parses all records in all sheets, and returns a list of dicts representing
        records found in sheets. From sheet to sheet records can contain different sets of fields, which is allowed.
        Parsing is done with ImportSheetParser class, with using field and row tags that have to be present in
        MS Excel sheets.

        :param xlsdir: path to directory with MS Excel *.xlsx workbooks
        :param tolog: logger object to use
        :param pfile: file to record import progress, one single file created, by default is "progress" in the
                      same xlsdir folder, or full path/filename can be specified explicitly by this parameter
        :param loglist: if reference to the list object provided, then logging to this list by using append() method,
                        each record is a dict like {"level": log level, "msg": log message}
                        in case if loglist is not None logger provider by tolog parameter is not used
        :param logdebug: log debug messages
        :return: list of task records loaded from excel sheets
        """
        # List of workbooks
        wbs = dict()
        # Name of progress record file - it must be ignored by the import
        if pfile is None:
            _pfile = "progress"
        else:
            _pfile = pfile
        # Load all files from folder
        _log_wbs_list = list()
        for _tmp_root, _tmp_dirs, _tmp_files in walk(xlsdir):
            for _tmp_name_walk_f in _tmp_files:
                try:
                    if _tmp_name_walk_f != _pfile:
                        wbs[_tmp_name_walk_f] = load_workbook(filename=f"{xlsdir.strip('/')}/{_tmp_name_walk_f}")
                        _log_wbs_list.append(_tmp_name_walk_f)
                except Exception:
                    if isinstance(loglist, list):
                        loglist.append({"level": "WARNING",
                                       "msg": f"Can't load {xlsdir.strip('/')}/{_tmp_name_walk_f}:\n{format_exc()}"})
                    else:
                        tolog.warning(f"Can't load {xlsdir.strip('/')}/{_tmp_name_walk_f}:\n{format_exc()}")
        # Log loaded workbooks
        if _log_wbs_list:
            _log_str = (f"\nLoaded {len(_log_wbs_list)} "
                        f"MS Excel workbook{'s' if len(_log_wbs_list) != 1 else ''} from "
                        f"file{'s' if len(_log_wbs_list) > 1 else ''}: "
                        f"{', '.join([(chr(39) + _i + chr(39)) for _i in _log_wbs_list])}")
        else:
            _log_str = "\nNo MS Excel workbooks loaded"
        if isinstance(loglist, list):
            loglist.append({"level": "INFO",
                            "msg": _log_str})
        else:
            tolog.info(_log_str)
        # Full list for importing
        reclist = list()
        # Log record string
        _log_str_loading = "\nSummary:\n"
        # Run checks, append correct records to the import list
        for _wname, _w in wbs.items():
            _sheetnames = _w.sheetnames
            _loaded_count = 0
            _log_str_loading += f"   '{_wname}':\n"
            for _sname in _sheetnames:
                try:
                    _sheet_parser = ImportSheetParser(_w[_sname], _sname)
                    if not _sheet_parser.ok():
                        raise ImportSheetParserException(f"      ERROR: failed loading sheet '{_sname}'")
                    _parsed_records = _sheet_parser.getparsedrecords()
                    for _r in _parsed_records:
                        reclist.append({**_r, "_wname": _wname, "_sname": _sname})
                    _log_str_loading += _sheet_parser.getlog()
                    _log_str_loading += (f"      TOTAL:  parsed records {_sheet_parser.getpasrsestats()['OK']} "
                                         f"of {_sheet_parser.getpasrsestats()['TOTAL']}\n")
                    _loaded_count += 1
                except ImportSheetParserException as _e:
                    _log_str_loading += f"      {repr(_e)}\n"
                except ImportCellParserException as _e:
                    _log_str_loading += f"      {repr(_e)}\n"
            _log_str_loading += (f"      loaded {_loaded_count} sheet{'s' if _loaded_count != 1 else ''} "
                                 f"out of {len(_sheetnames)}\n")
            # And finally close the workbook
            _w.close()
        _log_str_loading += (f"   TOTAL {len(reclist)} "
                             f"record{'s' if len(reclist) != 1 else ''}\n")
        # Log records loading
        if isinstance(loglist, list):
            loglist.append({"level": "INFO",
                            "msg": _log_str_loading})
        else:
            tolog.info(_log_str_loading)
        # Log full imported list
        _log_str_reclist = str()
        for _r in reclist:
            _log_str_reclist += str(_r) + "\n"
        if logdebug:
            if isinstance(loglist, list):
                loglist.append({"level": "DEBUG",
                                "msg": _log_str_reclist})
            else:
                tolog.debug(_log_str_reclist)
        # And finally return the list
        return reclist

    @staticmethod
    def jsondef(obj: Any) -> str:
        """
        :param obj: object unknown to json.dump
        :return: serialized value of object unknown to json.dump
        """
        if isinstance(obj, timedelta):
            return str(obj)
        elif isinstance(obj, datetime):
            return obj.strftime("%d.%m.%Y")
        return repr(obj)
